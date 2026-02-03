#!/usr/bin/env python3
"""Debug script to inspect Excel image anchors."""

import sys
from pathlib import Path
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))

def inspect_images(excel_path: str):
    """Inspect all image anchors in the Excel file."""
    wb = load_workbook(excel_path, data_only=True)
    sheet = wb.active
    
    print(f"Sheet: {sheet.title}")
    print(f"Max rows: {sheet.max_row}")
    
    if not hasattr(sheet, '_images') or not sheet._images:
        print("No images found")
        return
    
    print(f"\nTotal images: {len(sheet._images)}")
    print("="*100)
    
    for idx, img in enumerate(sheet._images[:3], 1):
        print(f"\nImage {idx}:")
        
        if not hasattr(img, 'anchor'):
            print("  No anchor property")
            continue
        
        anchor = img.anchor
        anchor_type = type(anchor).__name__
        print(f"  Anchor type: {anchor_type}")
        
        print(f"  All attributes: {[attr for attr in dir(anchor) if not attr.startswith('_')]}")
        
        if hasattr(anchor, '_from'):
            from_marker = anchor._from
            print(f"  _from.row: {from_marker.row} (1-indexed: {from_marker.row + 1})")
            print(f"  _from.col: {from_marker.col} (1-indexed: {from_marker.col + 1})")
            
            if hasattr(from_marker, 'rowOff'):
                print(f"  _from.rowOff: {from_marker.rowOff} EMUs")
            if hasattr(from_marker, 'colOff'):
                print(f"  _from.colOff: {from_marker.colOff} EMUs")
        
        try:
            if hasattr(anchor, '_to'):
                to_marker = anchor._to
                print(f"  _to.row: {to_marker.row} (1-indexed: {to_marker.row + 1})")
                print(f"  _to.col: {to_marker.col} (1-indexed: {to_marker.col + 1})")
            else:
                print("  No _to property")
        except Exception as e:
            print(f"  Error accessing _to: {e}")
        
        try:
            if hasattr(anchor, 'ext'):
                print(f"  ext.cx: {anchor.ext.cx}")
                print(f"  ext.cy: {anchor.ext.cy}")
            else:
                print("  No ext property")
        except Exception as e:
            print(f"  Error accessing ext: {e}")
        
        try:
            img_data = img._data()
            from PIL import Image as PILImage
            import io
            pil_img = PILImage.open(io.BytesIO(img_data))
            print(f"  Image dimensions: {pil_img.size[0]}x{pil_img.size[1]} pixels")
            
            typical_row_height_px = 20
            approx_row_span = max(1, round(pil_img.size[1] / typical_row_height_px))
            print(f"  Estimated row span: ~{approx_row_span} rows")
            
        except Exception as e:
            print(f"  Error getting image dimensions: {e}")
        
        print("-"*100)

if __name__ == "__main__":
    excel_file = "docs/1947688(ED736A0501)-case.xlsx"
    inspect_images(excel_file)
