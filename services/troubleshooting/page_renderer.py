#!/usr/bin/env python3
"""
Page Renderer for troubleshooting Excel files.

Converts Excel to PDF using LibreOffice headless and renders page images.
Also builds page context mapping rows/images to pages based on Excel page breaks.
"""

import logging
import os
import shutil
import subprocess
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from pdf2image import convert_from_path

logger = logging.getLogger(__name__)


@dataclass
class PageRenderResult:
    """Result of page rendering and context mapping."""

    pdf_path: Path
    page_images: List[Path]
    page_context: Dict[str, Dict]
    page_ranges: List[Tuple[int, int]]


class PageRenderer:
    """Render Excel pages and build per-page context."""

    def __init__(
        self,
        output_dir: Path,
        dpi: Optional[int] = None,
        libreoffice_path: Optional[str] = None,
        rows_per_page_fallback: Optional[int] = None
    ) -> None:
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.dpi = dpi or int(os.getenv("VLM_PAGE_RENDER_DPI", "150"))
        self.libreoffice_path = libreoffice_path or os.getenv("LIBREOFFICE_PATH", "libreoffice")
        self.rows_per_page_fallback = rows_per_page_fallback or int(os.getenv("VLM_ROWS_PER_PAGE", "50"))

    def render(self, excel_path: Path, case_data: Dict) -> PageRenderResult:
        """
        Render Excel to page images and build page context.

        Args:
            excel_path: Excel file to render
            case_data: Extracted case data (used for row/image mapping)

        Returns:
            PageRenderResult with PDF path, page images, context, and ranges
        """
        pdf_path = self._convert_excel_to_pdf(excel_path)
        page_images = self._convert_pdf_to_images(pdf_path)
        page_ranges = self._get_page_ranges(excel_path)
        page_context = self._build_page_context(case_data, page_ranges)
        return PageRenderResult(
            pdf_path=pdf_path,
            page_images=page_images,
            page_context=page_context,
            page_ranges=page_ranges
        )

    def _convert_excel_to_pdf(self, excel_path: Path) -> Path:
        """Convert Excel to PDF using LibreOffice headless."""
        pdf_dir = self.output_dir / "pdf"
        pdf_dir.mkdir(parents=True, exist_ok=True)

        libreoffice = self._resolve_libreoffice()
        if not libreoffice:
            raise FileNotFoundError(
                "LibreOffice not found. Install LibreOffice or set LIBREOFFICE_PATH to the executable."
            )

        command = [
            libreoffice,
            "--headless",
            "--convert-to",
            "pdf",
            "--outdir",
            str(pdf_dir),
            str(excel_path)
        ]

        logger.info("Rendering Excel to PDF via LibreOffice")
        try:
            subprocess.run(command, check=True, capture_output=True)
        except subprocess.CalledProcessError as exc:
            stderr = exc.stderr.decode("utf-8", errors="ignore") if exc.stderr else ""
            raise RuntimeError(f"LibreOffice conversion failed: {stderr}") from exc

        pdf_path = pdf_dir / f"{excel_path.stem}.pdf"
        if not pdf_path.exists():
            raise FileNotFoundError(f"PDF not found after conversion: {pdf_path}")
        return pdf_path

    def _resolve_libreoffice(self) -> Optional[str]:
        """Resolve LibreOffice executable path."""
        if self.libreoffice_path and shutil.which(self.libreoffice_path):
            return self.libreoffice_path
        for candidate in ("libreoffice", "soffice"):
            found = shutil.which(candidate)
            if found:
                return found
        return None

    def _convert_pdf_to_images(self, pdf_path: Path) -> List[Path]:
        """Convert PDF pages to PNG images."""
        pages_dir = self.output_dir / "pages"
        pages_dir.mkdir(parents=True, exist_ok=True)

        logger.info("Rendering PDF pages to images")
        images = convert_from_path(str(pdf_path), dpi=self.dpi)
        page_paths: List[Path] = []
        for index, image in enumerate(images, start=1):
            page_path = pages_dir / f"page_{index:02d}.png"
            image.save(page_path, "PNG")
            page_paths.append(page_path)

        return page_paths

    def _get_page_ranges(self, excel_path: Path) -> List[Tuple[int, int]]:
        """Determine row ranges per page based on Excel page breaks."""
        wb = load_workbook(excel_path, data_only=True)
        sheet = wb.active

        row_breaks = []
        if sheet.row_breaks:
            for brk in sheet.row_breaks:
                brk_id = getattr(brk, "id", None)
                if brk_id is None and isinstance(brk, tuple) and brk:
                    brk_id = brk[0]
                if brk_id:
                    try:
                        row_breaks.append(int(brk_id))
                    except (TypeError, ValueError):
                        continue

        row_breaks = sorted(set(row_breaks))

        ranges: List[Tuple[int, int]] = []
        if row_breaks:
            start = 1
            for brk in row_breaks:
                end = max(brk - 1, start)
                ranges.append((start, end))
                start = brk
            ranges.append((start, sheet.max_row))
        else:
            rows_per_page = self.rows_per_page_fallback
            start = 1
            while start <= sheet.max_row:
                end = min(start + rows_per_page - 1, sheet.max_row)
                ranges.append((start, end))
                start = end + 1

        return ranges

    def _build_page_context(self, case_data: Dict, page_ranges: List[Tuple[int, int]]) -> Dict[str, Dict]:
        """Build context map for rows and images on each page."""
        page_context: Dict[str, Dict] = {}

        def get_page_for_row(row_num: Optional[int]) -> Optional[int]:
            if row_num is None:
                return None
            for index, (start, end) in enumerate(page_ranges, start=1):
                if start <= row_num <= end:
                    return index
            return None

        for page_index, _ in enumerate(page_ranges, start=1):
            page_context[f"page_{page_index}"] = {
                "rows": [],
                "images": [],
                "continued_from_previous": False,
                "continues_to_next": False
            }

        for issue in case_data.get("issues", []):
            row_num = issue.get("excel_row")
            page_num = get_page_for_row(row_num)
            if page_num:
                page_context[f"page_{page_num}"]["rows"].append(issue.get("row_id"))

            for image in issue.get("images", []):
                anchor = image.get("anchor", {})
                anchor_row = anchor.get("row")
                image_page = get_page_for_row(anchor_row)
                if image_page:
                    anchor["page"] = image_page
                    page_context[f"page_{image_page}"]["images"].append(image.get("image_id"))

        return page_context
