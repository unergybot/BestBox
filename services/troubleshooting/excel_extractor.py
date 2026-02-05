#!/usr/bin/env python3
"""
Excel Troubleshooting Case Extractor

Extracts structured data and embedded images from troubleshooting Excel files.
Handles complex Excel layouts with metadata sections and data tables.

Usage:
    from services.troubleshooting.excel_extractor import ExcelTroubleshootingExtractor

    extractor = ExcelTroubleshootingExtractor(output_dir=Path("data/troubleshooting/processed"))
    case_data = extractor.extract_case(excel_path)
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
import io
import json
import uuid
from pathlib import Path
from typing import Dict, List, Optional
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelTroubleshootingExtractor:
    """Extract structured data and images from troubleshooting Excel files"""

    def __init__(self, output_dir: Path):
        """
        Initialize extractor.

        Args:
            output_dir: Directory for processed output (JSON + images)
        """
        self.output_dir = Path(output_dir)
        self.images_dir = self.output_dir / "images"
        self.images_dir.mkdir(parents=True, exist_ok=True)

        logger.info(f"ExcelExtractor initialized: output={self.output_dir}")

    def extract_case(self, excel_path: Path) -> Dict:
        """
        Extract complete troubleshooting case from Excel file.

        Args:
            excel_path: Path to Excel file

        Returns:
            dict with case metadata, issues, and images
        """
        excel_path = Path(excel_path)

        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        logger.info(f"📄 Processing: {excel_path.name}")

        try:
            # Load workbook with openpyxl for images and metadata
            wb = load_workbook(excel_path, data_only=True)
            sheet = wb.active

            logger.info(f"   Sheet: {sheet.title}, Rows: {sheet.max_row}, Cols: {sheet.max_column}")

            # Step 1: Extract metadata from header rows
            metadata = self._extract_metadata(sheet, excel_path)

            # Step 2: Extract data table (find header row automatically)
            header_row = self._find_data_table_header(sheet)
            logger.info(f"   Data table starts at row {header_row + 1}")

            df = pd.read_excel(excel_path, sheet_name=0, header=header_row)
            df_clean = df.dropna(subset=['NO', '問題点'], how='all')

            # Remove duplicate header rows that sometimes appear in data
            df_clean = df_clean[df_clean['NO'] != 'NO']

            # Step 3: Extract and map images
            image_map = self._extract_images(sheet, excel_path.stem)
            logger.info(f"   Extracted {len(image_map)} images")

            # Step 4: Build issues list with mapped images
            issues = self._build_issues(df_clean, image_map, metadata, header_row)
            logger.info(f"   Found {len(issues)} troubleshooting issues")

            # Step 5: Build case structure
            case_id = self._generate_case_id(metadata)

            case = {
                "case_id": case_id,
                "metadata": metadata,
                "issues": issues,
                "total_issues": len(issues),
                "source_file": str(excel_path),
                "extraction_version": "1.0",
                "vlm_validation": {
                    "status": "not_started",
                    "validated_at": None,
                    "pages_processed": 0,
                    "total_images": sum(len(issue.get('images', [])) for issue in issues),
                    "auto_corrected": 0,
                    "pending_review": 0,
                    "average_confidence": 0.0
                }
            }

            # Save to JSON
            json_path = self.output_dir / f"{case_id}.json"
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(case, f, ensure_ascii=False, indent=2)

            logger.info(f"   ✅ Saved to {json_path.name}")

            return case

        except Exception as e:
            logger.error(f"   ❌ Error processing {excel_path.name}: {e}")
            raise

    def _extract_metadata(self, sheet, excel_path: Path) -> Dict:
        """Extract case metadata from header rows (1-19)"""

        def safe_get(cell_ref):
            """Safely get cell value"""
            try:
                val = sheet[cell_ref].value
                return str(val).strip() if val not in (None, '') else None
            except:
                return None

        metadata = {
            "part_number": safe_get('F6'),
            "internal_number": safe_get('F8'),
            "mold_type": safe_get('F4'),
            "material_t0": safe_get('G13'),
            "material_t1": safe_get('I13'),
            "material_t2": safe_get('K13'),
            "color": safe_get('G14'),
            "molding_machine": safe_get('G19'),
            "source_filename": excel_path.name
        }

        return metadata

    def _find_data_table_header(self, sheet) -> int:
        """
        Find the row where the data table starts by looking for column headers.

        Returns:
            Row index (0-based for pandas) where data table header is located
        """
        # Look for the row containing key column headers
        key_headers = ['NO', '問題点', '原因，对策', '型试']

        for row_idx in range(15, 30):  # Check rows 15-30
            row_values = []
            for col_idx in range(1, 15):  # Check first 14 columns
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    row_values.append(str(cell_value))

            # Check if this row contains the key headers
            matches = sum(1 for header in key_headers if any(header in val for val in row_values))

            if matches >= 3:  # At least 3 key headers found
                return row_idx - 1  # Return 0-indexed row for pandas

        # Default to row 19 (0-indexed) if not found
        logger.warning("Could not auto-detect data table header, using default row 19")
        return 19

    def _extract_images(self, sheet, case_stem: str) -> Dict:
        """
        Extract all embedded images from sheet and save to files.
        Captures full cell range extent (top-left to bottom-right) for accurate
        image-to-issue mapping.

        Args:
            sheet: openpyxl worksheet
            case_stem: Case filename stem for naming images

        Returns:
            dict mapping image index to {image_id, file_path, anchor}
                where anchor contains {row_start, row_end, col_start, col_end, offsets}
        """
        image_map = {}

        if not hasattr(sheet, '_images') or not sheet._images:
            return image_map

        for idx, img in enumerate(sheet._images, 1):
            # Skip first image (usually company logo in header)
            if idx == 1:
                continue

            # Initialize anchor data with complete cell range information
            anchor_data = {
                "row_start": 0,
                "row_end": 0,
                "col_start": 0,
                "col_end": 0,
                "row_offs_top": 0,      # Offset from top of row_start (EMU units)
                "row_offs_bottom": 0,   # Offset at row_end (EMU units)
                "col_offs_left": 0,     # Offset from left of col_start (EMU units)
                "col_offs_right": 0,    # Offset at col_end (EMU units)
                "height_emu": 0,        # Height in EMU units
                "width_emu": 0,         # Width in EMU units
                "anchor_type": "unknown"
            }

            if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                # Top-left corner (always available)
                anchor_data["row_start"] = img.anchor._from.row + 1  # 1-indexed
                anchor_data["col_start"] = img.anchor._from.col + 1
                anchor_data["anchor_type"] = "oneCell"

                # Get offsets if available (in EMUs - English Metric Units)
                if hasattr(img.anchor._from, 'rowOff'):
                    anchor_data["row_offs_top"] = img.anchor._from.rowOff
                if hasattr(img.anchor._from, 'colOff'):
                    anchor_data["col_offs_left"] = img.anchor._from.colOff

                # Check for TwoCellAnchor (has 'to' property - bottom-right explicitly defined)
                if hasattr(img.anchor, 'to'):
                    anchor_data["anchor_type"] = "twoCell"
                    anchor_data["row_end"] = img.anchor.to.row + 1
                    anchor_data["col_end"] = img.anchor.to.col + 1

                    if hasattr(img.anchor.to, 'rowOff'):
                        anchor_data["row_offs_bottom"] = img.anchor.to.rowOff
                    if hasattr(img.anchor.to, 'colOff'):
                        anchor_data["col_offs_right"] = img.anchor.to.colOff

                    # Calculate height/width from cell range
                    anchor_data["height_emu"] = (
                        (anchor_data["row_end"] - anchor_data["row_start"]) * 914400 +
                        anchor_data["row_offs_bottom"] - anchor_data["row_offs_top"]
                    )
                    anchor_data["width_emu"] = (
                        (anchor_data["col_end"] - anchor_data["col_start"]) * 914400 +
                        anchor_data["col_offs_right"] - anchor_data["col_offs_left"]
                    )

                elif hasattr(img.anchor, 'ext'):
                    # OneCellAnchor with extent - calculate end from height/width
                    anchor_data["anchor_type"] = "oneCell"
                    anchor_data["height_emu"] = img.anchor.ext.cy
                    anchor_data["width_emu"] = img.anchor.ext.cx

                    # Approximate row span from height (assuming ~15 rows per inch)
                    # 914400 EMUs = 1 inch, typical row height varies
                    height_inches = img.anchor.ext.cy / 914400
                    approx_rows = max(1, round(height_inches * 15))
                    anchor_data["row_end"] = anchor_data["row_start"] + approx_rows

                    # Width usually doesn't span columns in these documents
                    anchor_data["col_end"] = anchor_data["col_start"]

                else:
                    # Fallback: single cell, no extent info
                    anchor_data["row_end"] = anchor_data["row_start"]
                    anchor_data["col_end"] = anchor_data["col_start"]

            # Generate image ID and filename
            image_id = f"{case_stem}_img{idx:03d}"
            image_filename = f"{image_id}.jpg"
            image_path = self.images_dir / image_filename

            try:
                # Get image data and convert to PIL
                img_data = img._data()
                pil_image = Image.open(io.BytesIO(img_data))

                # Convert to RGB if needed
                if pil_image.mode != 'RGB':
                    pil_image = pil_image.convert('RGB')

                # Save as JPEG
                pil_image.save(image_path, "JPEG", quality=90)

                image_map[idx] = {
                    "image_id": image_id,
                    "file_path": str(image_path),
                    "anchor": anchor_data
                }

            except Exception as e:
                logger.warning(f"Failed to extract image {idx}: {e}")

        return image_map

    def _build_issues(self, df: pd.DataFrame, image_map: Dict, metadata: Dict, header_row: int) -> List[Dict]:
        """
        Build issues list from data table with mapped images.

        Args:
            df: Cleaned DataFrame with issue data
            image_map: Extracted images mapped to positions
            metadata: Case metadata

        Returns:
            List of issue dictionaries
        """
        issues = []
        issue_map = {}

        for idx, row in df.iterrows():
            try:
                issue_num = int(row['NO'])
            except (ValueError, TypeError):
                continue

            excel_row = header_row + 2 + idx

            issue = {
                "issue_number": issue_num,
                "trial_version": str(row['型试']) if pd.notna(row['型试']) else None,
                "category": str(row['项目']) if pd.notna(row['项目']) else None,
                "problem": str(row['問題点']).strip() if pd.notna(row['問題点']) else "",
                "solution": str(row['原因，对策']).strip() if pd.notna(row['原因，对策']) else "",
                "result_t1": str(row['修正結果T1']) if pd.notna(row['修正結果T1']) else None,
                "result_t2": str(row['修正結果T2']) if pd.notna(row['修正結果T2']) else None,
                "cause_classification": str(row['原因分类']) if pd.notna(row['原因分类']) and '原因分类' in row else None,
                "row_id": f"r{len(issues) + 1}",
                "excel_row": excel_row,
                "images": [],
                "image_mapping_status": {
                    "total": 0,
                    "validated": 0,
                    "pending_review": 0
                }
            }
            issue_map[excel_row] = len(issues)
            issues.append(issue)

        image_candidates = {}

        for img_idx, img_data in image_map.items():
            anchor = img_data.get('anchor', {})
            image_id = img_data['image_id']

            candidates = []
            for issue_idx, issue in enumerate(issues):
                excel_row = issue['excel_row']

                match_result = self._calculate_image_issue_match(excel_row, anchor)

                if match_result['is_match']:
                    candidates.append((issue_idx, match_result, img_data, anchor))

            if candidates:
                image_candidates[image_id] = candidates

        for image_id, candidates in image_candidates.items():
            candidates.sort(key=lambda x: (-x[1]['confidence'], x[1]['row_distance']))

            best_match = candidates[0]
            issue_idx, match_result, img_data, anchor = best_match

            img_row_start = anchor.get('row_start', 0)
            img_row_end = anchor.get('row_end', img_row_start)
            img_col_start = anchor.get('col_start', 0)
            img_col_end = anchor.get('col_end', img_col_start)

            issues[issue_idx]['images'].append({
                "image_id": img_data['image_id'],
                "file_path": img_data['file_path'],
                "cell_location": f"Row {img_row_start}-{img_row_end}, Col {img_col_start}",
                "anchor": {
                    "row_start": img_row_start,
                    "row_end": img_row_end,
                    "col_start": img_col_start,
                    "col_end": img_col_end,
                    "page": None
                },
                "spatial_match": {
                    "type": match_result['match_type'],
                    "confidence": match_result['confidence'],
                    "row_distance": match_result['row_distance']
                },
                "mapping_validation": {
                    "status": "pending",
                    "method": "anchor_based",
                    "confidence": match_result['confidence'],
                    "vlm_reason": None,
                    "validated_at": None,
                    "reviewed_by": None
                },
                "vl_description": None,
                "defect_type": None,
                "text_in_image": None
            })

        for issue in issues:
            issue["image_mapping_status"]["total"] = len(issue["images"])

        return issues

    def _calculate_image_issue_match(self, issue_row: int, anchor: Dict) -> Dict:
        """Calculate image-to-issue matching for Excel troubleshooting files.

        In these files, images are typically placed BELOW their associated issue
        (3-15 rows below) and span multiple rows. The issue row should be above
        or at the start of the image range for a match.
        """
        img_row_start = anchor.get('row_start', 0)
        img_row_end = anchor.get('row_end', img_row_start)
        row_offs_top = anchor.get('row_offs_top', 0)

        row_span = img_row_end - img_row_start

        # Distance from issue to image start (positive = issue is above image)
        distance_to_image_start = img_row_start - issue_row

        # Check if issue is within image span
        within_image_span = img_row_start <= issue_row <= img_row_end

        # Check if issue is immediately above the image (typical case)
        issue_above_image = issue_row < img_row_start

        # Calculate how much the image extends below the issue
        if issue_above_image:
            image_overlap = row_span - distance_to_image_start
        else:
            image_overlap = img_row_end - issue_row

        # Reject if issue is far below the image (belongs to different issue)
        if issue_row > img_row_end + 5:
            return {
                'is_match': False,
                'match_type': 'none',
                'confidence': 0.0,
                'row_distance': issue_row - img_row_end
            }

        # Best match: issue is immediately above the image (distance 0-15 rows)
        if issue_above_image:
            if distance_to_image_start <= 3:
                # Issue very close to image start - high confidence primary match
                confidence = min(1.0, 1.0 - (distance_to_image_start * 0.1))
                match_type = 'primary'
            elif distance_to_image_start <= 8:
                # Issue reasonably close to image - medium confidence
                confidence = max(0.6, 0.85 - (distance_to_image_start * 0.05))
                match_type = 'secondary'
            elif distance_to_image_start <= 50:
                # Issue further away but still potentially related (e.g. attached page)
                confidence = max(0.4, 0.65 - (distance_to_image_start * 0.01))
                match_type = 'tertiary'
            else:
                # Too far, likely belongs to a different issue
                return {
                    'is_match': False,
                    'match_type': 'none',
                    'confidence': 0.0,
                    'row_distance': distance_to_image_start
                }
        elif within_image_span:
            # Issue is inside the image's row range
            if row_span <= 3:
                # Small image covering few rows
                confidence = 0.85
                match_type = 'inline'
            else:
                # Large image spanning many rows - could be shared
                confidence = 0.70
                match_type = 'overlap'
        else:
            # Issue is below image but within 5 rows - might be related
            distance_below = issue_row - img_row_end
            confidence = max(0.0, 0.35 - (distance_below * 0.07))
            match_type = 'post_image'

        # Boost confidence if image has a top offset (meaning it's positioned
        # within the starting cell, not exactly at the boundary)
        if row_offs_top > 100000 and issue_above_image and distance_to_image_start <= 5:
            confidence = min(1.0, confidence + 0.05)

        return {
            'is_match': True,
            'match_type': match_type,
            'confidence': round(confidence, 2),
            'row_distance': distance_to_image_start if issue_above_image else 0
        }

    def _generate_case_id(self, metadata: Dict) -> str:
        """Generate unique case ID from metadata"""

        part_num = metadata.get('part_number') or 'UNKNOWN'
        internal_num = metadata.get('internal_number') or str(uuid.uuid4())[:8]

        case_id = f"TS-{part_num}-{internal_num}"

        return case_id


# Convenience function for quick extraction
def extract_troubleshooting_case(excel_path: str, output_dir: str = "data/troubleshooting/processed") -> Dict:
    """
    Quick extraction function.

    Args:
        excel_path: Path to Excel file
        output_dir: Output directory

    Returns:
        Extracted case data
    """
    extractor = ExcelTroubleshootingExtractor(Path(output_dir))
    return extractor.extract_case(Path(excel_path))


if __name__ == "__main__":
    # Test with sample file
    import sys

    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    else:
        excel_file = "docs/1947688(ED736A0501)-case.xlsx"

    print(f"Testing extraction with: {excel_file}")
    case = extract_troubleshooting_case(excel_file)
    print(f"\n✅ Extracted case: {case['case_id']}")
    print(f"   Issues: {case['total_issues']}")
    print(f"   Images: {sum(len(issue['images']) for issue in case['issues'])}")
